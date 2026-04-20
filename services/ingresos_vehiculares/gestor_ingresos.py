"""
Gestor de Ingresos Vehiculares - Create, Read, Update
"""

import logging
import base64
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Configuración de rutas
BASE_REGISTROS = Path('/Users/doriantituana/Desktop/DMT_Registros_Ingresos/DMT_Registros_Ingresos_Vehiculares')
MESES = {
    1: ('Enero', 'ENE'),
    2: ('Febrero', 'FEB'),
    3: ('Marzo', 'MAR'),
    4: ('Abril', 'ABR'),
    5: ('Mayo', 'MAY'),
    6: ('Junio', 'JUN'),
    7: ('Julio', 'JUL'),
    8: ('Agosto', 'AGO'),
    9: ('Septiembre', 'SEP'),
    10: ('Octubre', 'OCT'),
    11: ('Noviembre', 'NOV'),
    12: ('Diciembre', 'DIC'),
}


class GestorIngresosVehiculares:
    """Gestor de operaciones CRUD para ingresos vehiculares"""

    def __init__(self):
        self.base_path = BASE_REGISTROS

    def _generar_ticket(self, fecha: datetime) -> str:
        """
        Genera número de ticket secuencial para el día
        Formato: TICKET-[MES]-[DÍA]-[NÚMERO]
        """
        año = fecha.year
        mes_num = fecha.month
        día = fecha.day

        nombre_mes, abrev_mes = MESES[mes_num]
        mes_path = self.base_path / str(año) / nombre_mes

        # Obtener lista de carpetas del día
        día_path = mes_path / f"{día:02d}"
        if not día_path.exists():
            return f"TICKET-{abrev_mes}-{día:02d}-001"

        # Contar tickets existentes en el día
        tickets_existentes = [d.name for d in día_path.iterdir() if d.is_dir() and d.name.startswith('TICKET')]
        numero_siguiente = len(tickets_existentes) + 1

        return f"TICKET-{abrev_mes}-{día:02d}-{numero_siguiente:03d}"

    def _crear_ticket_folder(self, ticket: str, fecha: datetime) -> Path:
        """Crea la carpeta del ticket con estructura TICKET-xxx/"""
        año = fecha.year
        mes_num = fecha.month
        día = fecha.day

        nombre_mes, _ = MESES[mes_num]

        # Crear ruta: año/mes/día/TICKET-xxx
        ticket_path = self.base_path / str(año) / nombre_mes / f"{día:02d}" / ticket
        ticket_path.mkdir(parents=True, exist_ok=True)

        return ticket_path

    def _validar_base64_imagen(self, imagen_base64: str, nombre_imagen: str) -> tuple[bool, str]:
        """
        Valida que el string sea un base64 válido decodificable.
        Retorna (válido, mensaje_error)
        """
        try:
            if not imagen_base64:
                return False, f"{nombre_imagen} no puede estar vacía"
            
            # Intentar decodificar
            decoded = base64.b64decode(imagen_base64)
            
            if len(decoded) == 0:
                return False, f"{nombre_imagen} decodificada está vacía"
            
            # Verificar que sea una imagen JPEG válida (empieza con FFD8)
            if not decoded.startswith(b'\xff\xd8'):
                return False, f"{nombre_imagen} no es una imagen JPEG válida"
            
            return True, ""
        except Exception as e:
            return False, f"Error decodificando {nombre_imagen}: {str(e)}"

    def _guardar_imagenes(self, ticket_path: Path, imagen_usuario_base64: str, 
                         imagen_cedula_base64: str, imagen_placa_base64: str) -> bool:
        """Guarda 3 imágenes en formato JPEG: usuario, cedula, placa"""
        try:
            # Guardar imagen de usuario
            usuario_bytes = base64.b64decode(imagen_usuario_base64)
            usuario_path = ticket_path / "usuario.jpeg"
            usuario_path.write_bytes(usuario_bytes)

            # Guardar imagen de cédula
            cedula_bytes = base64.b64decode(imagen_cedula_base64)
            cedula_path = ticket_path / "cedula.jpeg"
            cedula_path.write_bytes(cedula_bytes)

            # Guardar imagen de placa
            placa_bytes = base64.b64decode(imagen_placa_base64)
            placa_path = ticket_path / "placa.jpeg"
            placa_path.write_bytes(placa_bytes)

            logger.info(f"Imágenes guardadas en {ticket_path}")
            return True
        except Exception as e:
            logger.error(f"Error guardando imágenes: {e}")
            return False

    def _obtener_ruta_excel(self, fecha: datetime) -> Path:
        """Obtiene la ruta del Excel mensual"""
        año = fecha.year
        mes_num = fecha.month
        nombre_mes, _ = MESES[mes_num]

        excel_path = self.base_path / str(año) / nombre_mes / f"Registros_{nombre_mes}_{año}.xlsx"
        return excel_path

    def _crear_o_actualizar_excel(self, excel_path: Path, datos: dict) -> bool:
        """Crea o actualiza el Excel del mes"""
        try:
            # Verificar si Excel existe
            if excel_path.exists():
                wb = load_workbook(excel_path)
                ws = wb.active
            else:
                # Crear nuevo Excel con encabezados
                wb = Workbook()
                ws = wb.active
                ws.title = "Registros"

                # Encabezados - FORMATO CORRECTO (9 columnas - sin placa)
                headers = [
                    'Ticket', 'Fecha de Ingreso', 'Cédula', 'Nombres', 'Apellidos',
                    'Departamento', 'Motivo', 'Hora entrada', 'Hora salida'
                ]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            # Agregar fila con datos
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=datos['ticket'])
            ws.cell(row=next_row, column=2, value=datos['fecha_ingreso_fmt'])  # Fecha formateada dd/mm/yyyy
            ws.cell(row=next_row, column=3, value=datos['numero_cedula'])
            ws.cell(row=next_row, column=4, value=datos['nombre'])
            ws.cell(row=next_row, column=5, value=datos['apellido'])
            ws.cell(row=next_row, column=6, value=datos['departamento'])
            ws.cell(row=next_row, column=7, value=datos['motivo'])
            ws.cell(row=next_row, column=8, value=datos['hora_entrada'])
            ws.cell(row=next_row, column=9, value=datos['salida_estado'])  # Vacío

            wb.save(excel_path)
            logger.info(f"Excel actualizado: {excel_path}")
            return True
        except Exception as e:
            logger.error(f"Error actualizando Excel: {e}")
            return False

    def crear_ingreso(
        self,
        numero_cedula: str,
        nombres: str,
        apellidos: str,
        hora_entrada: str,
        departamento: str,
        motivo: str,
        imagen_usuario_base64: str,
        imagen_cedula_base64: str,
        imagen_placa_base64: str,
    ) -> dict:
        """
        Crea un nuevo registro de ingreso vehicular.
        PATRÓN TRANSACCIONAL: Valida TODO primero, solo crea si todo es válido.
        """
        try:
            # ===== FASE 1: VALIDACIÓN (Sin cambios en el sistema) =====
            
            # Validar datos básicos
            if not numero_cedula or not str(numero_cedula).strip():
                return {"error": "Cédula es requerida"}
            if not nombres or not str(nombres).strip():
                return {"error": "Nombres son requeridos"}
            if not apellidos or not str(apellidos).strip():
                return {"error": "Apellidos son requeridos"}
            if not hora_entrada or not str(hora_entrada).strip():
                return {"error": "Hora de entrada es requerida"}
            if not departamento or not str(departamento).strip():
                return {"error": "Departamento es requerido"}
            if not motivo or not str(motivo).strip():
                return {"error": "Motivo es requerido"}

            # Validar imágenes
            valido_usuario, msg_usuario = self._validar_base64_imagen(imagen_usuario_base64, "Imagen de usuario")
            if not valido_usuario:
                return {"error": msg_usuario}

            valido_cedula, msg_cedula = self._validar_base64_imagen(imagen_cedula_base64, "Imagen de cédula")
            if not valido_cedula:
                return {"error": msg_cedula}

            valido_placa, msg_placa = self._validar_base64_imagen(imagen_placa_base64, "Imagen de placa")
            if not valido_placa:
                return {"error": msg_placa}

            # ===== FASE 2: EJECUCIÓN (Solo si validación pasó) =====
            
            fecha_ahora = datetime.now()

            # Generar ticket
            ticket = self._generar_ticket(fecha_ahora)

            # Crear carpeta del ticket
            ticket_path = self._crear_ticket_folder(ticket, fecha_ahora)

            # Guardar imágenes (3 imágenes)
            if not self._guardar_imagenes(ticket_path, imagen_usuario_base64, imagen_cedula_base64, imagen_placa_base64):
                return {"error": "Error guardando imágenes"}

            # Preparar datos para Excel
            fecha_registro_str = fecha_ahora.strftime("%d/%m/%Y")
            datos_excel = {
                'ticket': ticket,
                'nombre': nombres,
                'apellido': apellidos,
                'numero_cedula': numero_cedula,
                'departamento': departamento,
                'motivo': motivo,
                'hora_entrada': hora_entrada,
                'salida_estado': '',  # Vacío al crear
                'fecha_ingreso_fmt': fecha_registro_str,  # dd/mm/yyyy
            }

            # Actualizar Excel
            excel_path = self._obtener_ruta_excel(fecha_ahora)
            if not self._crear_o_actualizar_excel(excel_path, datos_excel):
                return {"error": "Error actualizando Excel"}

            return {
                "exito": True,
                "ticket": ticket,
                "fecha_registro": fecha_registro_str,
                "hora_entrada": hora_entrada,
            }

        except Exception as e:
            logger.error(f"Error creando ingreso: {e}")
            return {"error": str(e)}

    def leer_ingreso(self, ticket: str) -> dict:
        """
        Lee un registro de ingreso por ticket
        Retorna todos los datos incluyendo las 3 imágenes en Base64
        """
        try:
            # Buscar ticket en todas las carpetas
            for año_path in self.base_path.iterdir():
                if not año_path.is_dir():
                    continue

                for mes_path in año_path.iterdir():
                    if not mes_path.is_dir():
                        continue

                    for día_path in mes_path.iterdir():
                        if not día_path.is_dir():
                            continue

                        ticket_path = día_path / ticket
                        if ticket_path.exists():
                            return self._extraer_datos_ticket(ticket_path, ticket, mes_path)

            return {"error": f"Ticket {ticket} no encontrado"}

        except Exception as e:
            logger.error(f"Error leyendo ingreso: {e}")
            return {"error": str(e)}

    def _extraer_datos_ticket(self, ticket_path: Path, ticket: str, mes_path: Path) -> dict:
        """Extrae datos del ticket incluyendo 3 imágenes en Base64"""
        try:
            # Leer imágenes (3)
            usuario_b64 = ""
            cedula_b64 = ""
            placa_b64 = ""

            usuario_path = ticket_path / "usuario.jpeg"
            if usuario_path.exists():
                usuario_b64 = base64.b64encode(usuario_path.read_bytes()).decode('utf-8')

            cedula_path = ticket_path / "cedula.jpeg"
            if cedula_path.exists():
                cedula_b64 = base64.b64encode(cedula_path.read_bytes()).decode('utf-8')

            placa_path = ticket_path / "placa.jpeg"
            if placa_path.exists():
                placa_b64 = base64.b64encode(placa_path.read_bytes()).decode('utf-8')

            # Leer datos del Excel del mes
            año = mes_path.parent.name
            mes_nombre = mes_path.name
            excel_path = mes_path / f"Registros_{mes_nombre}_{año}.xlsx"

            logger.info(f"Buscando Excel en: {excel_path}")

            if excel_path.exists():
                wb = load_workbook(excel_path)
                ws = wb.active

                # Buscar fila con el ticket (10 columnas)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == ticket:  # Columna 0: Ticket
                        return {
                            "exito": True,
                            "ticket": row[0],              # Columna 0: Ticket
                            "fecha_ingreso": row[1],       # Columna 1: Fecha de Ingreso
                            "numero_cedula": row[2],       # Columna 2: Cédula
                            "nombres": row[3],             # Columna 3: Nombres
                            "apellidos": row[4],           # Columna 4: Apellidos
                            "departamento": row[5],        # Columna 5: Departamento
                            "motivo": row[6],              # Columna 6: Motivo
                            "hora_entrada": row[7],        # Columna 7: Hora entrada
                            "hora_salida": row[8],         # Columna 8: Hora salida
                            "imagen_usuario_base64": usuario_b64,
                            "imagen_cedula_base64": cedula_b64,
                            "imagen_placa_base64": placa_b64,
                        }
                
                logger.warning(f"Ticket {ticket} no encontrado en Excel {excel_path}")
            else:
                logger.warning(f"Excel no encontrado en: {excel_path}")

            # Si no se encontró en Excel, retornar datos básicos
            return {
                "exito": True,
                "ticket": ticket,
                "imagen_usuario_base64": usuario_b64,
                "imagen_cedula_base64": cedula_b64,
                "imagen_placa_base64": placa_b64,
                "nota": "Datos de imágenes sin información de Excel",
            }

        except Exception as e:
            logger.error(f"Error extrayendo datos: {e}")
            return {"error": str(e)}

    def actualizar_ingreso(self, ticket: str, datos_actualizacion: dict) -> dict:
        """
        Actualiza un registro de ingreso (solo ciertos campos)
        Campos permitidos: numero_cedula, nombres, apellidos, departamento, motivo
        La fecha NO se modifica. La placa solo se guarda como imagen, no se edita.
        """
        try:
            # Buscar el Excel que contiene el ticket
            for año_path in self.base_path.iterdir():
                if not año_path.is_dir():
                    continue

                for mes_path in año_path.iterdir():
                    if not mes_path.is_dir():
                        continue

                    # Construir ruta del Excel
                    excel_path = mes_path / f"Registros_{mes_path.name}_{año_path.name}.xlsx"

                    if excel_path.exists():
                        wb = load_workbook(excel_path)
                        ws = wb.active

                        # Buscar fila con el ticket (9 columnas)
                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                            if row[0].value == ticket:  # Columna 0: Ticket
                                # Actualizar solo campos permitidos
                                if 'numero_cedula' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=3, value=datos_actualizacion['numero_cedula'])  # Columna 3: Cédula
                                if 'nombres' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=4, value=datos_actualizacion['nombres'])  # Columna 4: Nombres
                                if 'apellidos' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=5, value=datos_actualizacion['apellidos'])  # Columna 5: Apellidos
                                if 'departamento' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=6, value=datos_actualizacion['departamento'])  # Columna 6: Departamento
                                if 'motivo' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=7, value=datos_actualizacion['motivo'])  # Columna 7: Motivo

                                wb.save(excel_path)
                                logger.info(f"Ticket {ticket} actualizado")
                                return {"exito": True, "ticket": ticket, "mensaje": "Registro actualizado exitosamente"}

            return {"error": f"Ticket {ticket} no encontrado"}

        except Exception as e:
            logger.error(f"Error actualizando ingreso: {e}")
            return {"error": str(e)}

    def listar_ingresos_por_dia(self, fecha_str: str) -> dict:
        """
        Lista todos los tickets de un día específico
        fecha_str: formato DD/MM/YYYY o YYYY-MM-DD
        Retorna: lista de registros del día (sin imágenes Base64)
        """
        try:
            # Parsear fecha
            try:
                if '/' in fecha_str:
                    fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
                else:
                    fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
            except ValueError:
                return {"error": "Formato fecha inválido. Use DD/MM/YYYY o YYYY-MM-DD"}

            año = fecha.year
            mes_num = fecha.month
            día = fecha.day

            nombre_mes, _ = MESES[mes_num]

            # Ruta al Excel del mes
            mes_path = self.base_path / str(año) / nombre_mes
            excel_path = mes_path / f"Registros_{nombre_mes}_{año}.xlsx"

            if not excel_path.exists():
                return {
                    "exito": True,
                    "fecha": fecha_str,
                    "cantidad_tickets": 0,
                    "tickets": []
                }

            # Abrir Excel y filtrar por día
            wb = load_workbook(excel_path)
            ws = wb.active

            tickets_día = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # Saltar filas vacías
                    continue

                # row[1] es la fecha en formato dd/mm/yyyy
                try:
                    if row[1]:
                        fecha_fila = datetime.strptime(str(row[1]), "%d/%m/%Y")
                        if fecha_fila.day == día and fecha_fila.month == mes_num and fecha_fila.year == año:
                            tickets_día.append({
                                "ticket": row[0],              # Columna 0: Ticket
                                "fecha_ingreso": row[1],       # Columna 1: Fecha de Ingreso
                                "numero_cedula": row[2],       # Columna 2: Cédula
                                "nombres": row[3],             # Columna 3: Nombres
                                "apellidos": row[4],           # Columna 4: Apellidos
                                "departamento": row[5],        # Columna 5: Departamento
                                "motivo": row[6],              # Columna 6: Motivo
                                "hora_entrada": row[7],        # Columna 7: Hora entrada
                                "hora_salida": row[8],         # Columna 8: Hora salida
                            })
                except Exception as e:
                    logger.warning(f"Error procesando fila con ticket {row[0]}: {e}")
                    continue

            return {
                "exito": True,
                "fecha": fecha_str,
                "cantidad_tickets": len(tickets_día),
                "tickets": tickets_día
            }

        except Exception as e:
            logger.error(f"Error listando ingresos por día: {e}")
            return {"error": str(e)}
