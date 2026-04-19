"""
Gestor de Ingresos Peatonales - Create, Read, Update
"""

import logging
import base64
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Configuración de rutas
BASE_REGISTROS = Path('/Users/doriantituana/Desktop/DMT_Registros_Ingresos/DMT_Registros_Ingresos_Peatonales')
MESES = {
    1: ('Enero', 'ENE'),
    2: ('Febrero', 'FEB'),
    3: ('Marzo', 'MAR'),
    4: ('Abril', 'ABR'),
    5: ('Mayo', 'MAY'),
    6: ('Junio', 'JUN'),
    7: ('Julio', 'JUL'),
    8: ('Agosto', 'AGO'),
    9: ('Septiembre', 'SEP'),
    10: ('Octubre', 'OCT'),
    11: ('Noviembre', 'NOV'),
    12: ('Diciembre', 'DIC'),
}


class GestorIngresosPeatonales:
    """Gestor de operaciones CRUD para ingresos peatonales"""

    def __init__(self):
        self.base_path = BASE_REGISTROS

    def _generar_ticket(self, fecha: datetime) -> str:
        """
        Genera número de ticket secuencial para el día
        Formato: TICKET-[MES]-[DÍA]-[NÚMERO]
        """
        año = fecha.year
        mes_num = fecha.month
        día = fecha.day

        nombre_mes, abrev_mes = MESES[mes_num]
        mes_path = self.base_path / str(año) / nombre_mes

        # Obtener lista de carpetas del día
        día_path = mes_path / f"{día:02d}"
        if not día_path.exists():
            return f"TICKET-{abrev_mes}-{día:02d}-001"

        # Contar tickets existentes en el día
        tickets_existentes = [d.name for d in día_path.iterdir() if d.is_dir() and d.name.startswith('TICKET')]
        numero_siguiente = len(tickets_existentes) + 1

        return f"TICKET-{abrev_mes}-{día:02d}-{numero_siguiente:03d}"

    def _crear_ticket_folder(self, ticket: str, fecha: datetime) -> Path:
        """Crea la carpeta del ticket con estructura TICKET-xxx/"""
        año = fecha.year
        mes_num = fecha.month
        día = fecha.day

        nombre_mes, _ = MESES[mes_num]

        # Crear ruta: año/mes/día/TICKET-xxx
        ticket_path = self.base_path / str(año) / nombre_mes / f"{día:02d}" / ticket
        ticket_path.mkdir(parents=True, exist_ok=True)

        return ticket_path

    def _guardar_imagenes(self, ticket_path: Path, imagen_usuario_base64: str, imagen_cedula_base64: str) -> bool:
        """Guarda imágenes en formato JPEG"""
        try:
            # Guardar imagen de usuario
            usuario_bytes = base64.b64decode(imagen_usuario_base64)
            usuario_path = ticket_path / "usuario.jpeg"
            usuario_path.write_bytes(usuario_bytes)

            # Guardar imagen de cédula
            cedula_bytes = base64.b64decode(imagen_cedula_base64)
            cedula_path = ticket_path / "cedula.jpeg"
            cedula_path.write_bytes(cedula_bytes)

            logger.info(f"Imágenes guardadas en {ticket_path}")
            return True
        except Exception as e:
            logger.error(f"Error guardando imágenes: {e}")
            return False

    def _obtener_ruta_excel(self, fecha: datetime) -> Path:
        """Obtiene la ruta del Excel mensual"""
        año = fecha.year
        mes_num = fecha.month
        nombre_mes, _ = MESES[mes_num]

        excel_path = self.base_path / str(año) / nombre_mes / f"Registros_{nombre_mes}_{año}.xlsx"
        return excel_path

    def _crear_o_actualizar_excel(self, excel_path: Path, datos: dict) -> bool:
        """Crea o actualiza el Excel del mes"""
        try:
            # Verificar si Excel existe
            if excel_path.exists():
                wb = load_workbook(excel_path)
                ws = wb.active
            else:
                # Crear nuevo Excel con encabezados
                wb = Workbook()
                ws = wb.active
                ws.title = "Registros"

                # Encabezados - FORMATO CORRECTO
                headers = [
                    'Ticket', 'Fecha de Ingreso', 'Cédula', 'Nombres', 'Apellidos',
                    'Departamento', 'Motivo', 'Hora entrada', 'Hora salida'
                ]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            # Agregar fila con datos
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=datos['ticket'])
            ws.cell(row=next_row, column=2, value=datos['fecha_ingreso_fmt'])  # Fecha formateada dd/mm/yyyy
            ws.cell(row=next_row, column=3, value=datos['numero_cedula'])
            ws.cell(row=next_row, column=4, value=datos['nombre'])
            ws.cell(row=next_row, column=5, value=datos['apellido'])
            ws.cell(row=next_row, column=6, value=datos['departamento'])
            ws.cell(row=next_row, column=7, value=datos['motivo'])
            ws.cell(row=next_row, column=8, value=datos['hora_entrada'])
            ws.cell(row=next_row, column=9, value=datos['salida_estado'])  # Vacío

            wb.save(excel_path)
            logger.info(f"Excel actualizado: {excel_path}")
            return True
        except Exception as e:
            logger.error(f"Error actualizando Excel: {e}")
            return False

    def crear_ingreso(
        self,
        numero_cedula: str,
        nombres: str,
        apellidos: str,
        hora_entrada: str,
        departamento: str,
        motivo: str,
        imagen_usuario_base64: str,
        imagen_cedula_base64: str,
    ) -> dict:
        """
        Crea un nuevo registro de ingreso peatonal
        """
        try:
            fecha_ahora = datetime.now()

            # Generar ticket
            ticket = self._generar_ticket(fecha_ahora)

            # Crear carpeta del ticket
            ticket_path = self._crear_ticket_folder(ticket, fecha_ahora)

            # Guardar imágenes
            if not self._guardar_imagenes(ticket_path, imagen_usuario_base64, imagen_cedula_base64):
                return {"error": "Error guardando imágenes"}

            # Preparar datos para Excel
            fecha_registro_str = fecha_ahora.strftime("%d/%m/%Y")
            datos_excel = {
                'ticket': ticket,
                'nombre': nombres,
                'apellido': apellidos,
                'numero_cedula': numero_cedula,
                'departamento': departamento,
                'motivo': motivo,
                'hora_entrada': hora_entrada,
                'salida_estado': '',  # Vacío al crear
                'fecha_ingreso_fmt': fecha_registro_str,  # dd/mm/yyyy
            }

            # Actualizar Excel
            excel_path = self._obtener_ruta_excel(fecha_ahora)
            if not self._crear_o_actualizar_excel(excel_path, datos_excel):
                return {"error": "Error actualizando Excel"}

            return {
                "exito": True,
                "ticket": ticket,
                "fecha_registro": fecha_registro_str,
                "hora_entrada": hora_entrada,
            }

        except Exception as e:
            logger.error(f"Error creando ingreso: {e}")
            return {"error": str(e)}

    def leer_ingreso(self, ticket: str) -> dict:
        """
        Lee un registro de ingreso por ticket
        Retorna todos los datos incluyendo imágenes en Base64
        """
        try:
            # Buscar ticket en todas las carpetas
            for año_path in self.base_path.iterdir():
                if not año_path.is_dir():
                    continue

                for mes_path in año_path.iterdir():
                    if not mes_path.is_dir():
                        continue

                    for día_path in mes_path.iterdir():
                        if not día_path.is_dir():
                            continue

                        ticket_path = día_path / ticket
                        if ticket_path.exists():
                            return self._extraer_datos_ticket(ticket_path, ticket, mes_path)

            return {"error": f"Ticket {ticket} no encontrado"}

        except Exception as e:
            logger.error(f"Error leyendo ingreso: {e}")
            return {"error": str(e)}

    def _extraer_datos_ticket(self, ticket_path: Path, ticket: str, mes_path: Path) -> dict:
        """Extrae datos del ticket incluyendo imágenes en Base64"""
        try:
            # Leer imágenes
            usuario_b64 = ""
            cedula_b64 = ""

            usuario_path = ticket_path / "usuario.jpeg"
            if usuario_path.exists():
                usuario_b64 = base64.b64encode(usuario_path.read_bytes()).decode('utf-8')

            cedula_path = ticket_path / "cedula.jpeg"
            if cedula_path.exists():
                cedula_b64 = base64.b64encode(cedula_path.read_bytes()).decode('utf-8')

            # Leer datos del Excel del mes
            excel_path = mes_path / f"Registros_{mes_path.name}_{ticket_path.parent.parent.name}.xlsx"

            if excel_path.exists():
                wb = load_workbook(excel_path)
                ws = wb.active

                # Buscar fila con el ticket
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == ticket:  # Columna 0: Ticket
                        return {
                            "exito": True,
                            "ticket": row[0],              # Columna 0: Ticket
                            "fecha_ingreso": row[1],       # Columna 1: Fecha de Ingreso
                            "numero_cedula": row[2],       # Columna 2: Cédula
                            "nombres": row[3],             # Columna 3: Nombres
                            "apellidos": row[4],           # Columna 4: Apellidos
                            "departamento": row[5],        # Columna 5: Departamento
                            "motivo": row[6],              # Columna 6: Motivo
                            "hora_entrada": row[7],        # Columna 7: Hora entrada
                            "hora_salida": row[8],         # Columna 8: Hora salida
                            "imagen_usuario_base64": usuario_b64,
                            "imagen_cedula_base64": cedula_b64,
                        }

            # Si no se encontró en Excel, retornar datos básicos
            return {
                "exito": True,
                "ticket": ticket,
                "imagen_usuario_base64": usuario_b64,
                "imagen_cedula_base64": cedula_b64,
                "nota": "Datos de imágenes sin información de Excel",
            }

        except Exception as e:
            logger.error(f"Error extrayendo datos: {e}")
            return {"error": str(e)}

    def actualizar_ingreso(self, ticket: str, datos_actualizacion: dict) -> dict:
        """
        Actualiza un registro de ingreso (solo ciertos campos)
        Campos permitidos: numero_cedula, nombres, apellidos, departamento, motivo
        La fecha NO se modifica
        """
        try:
            # Buscar el Excel que contiene el ticket
            for año_path in self.base_path.iterdir():
                if not año_path.is_dir():
                    continue

                for mes_path in año_path.iterdir():
                    if not mes_path.is_dir():
                        continue

                    # Construir ruta del Excel
                    excel_path = mes_path / f"Registros_{mes_path.name}_{año_path.name}.xlsx"

                    if excel_path.exists():
                        wb = load_workbook(excel_path)
                        ws = wb.active

                        # Buscar fila con el ticket
                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                            if row[0].value == ticket:  # Columna 0: Ticket
                                # Actualizar solo campos permitidos (nuevas posiciones)
                                if 'numero_cedula' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=3, value=datos_actualizacion['numero_cedula'])  # Columna 3: Cédula
                                if 'nombres' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=4, value=datos_actualizacion['nombres'])  # Columna 4: Nombres
                                if 'apellidos' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=5, value=datos_actualizacion['apellidos'])  # Columna 5: Apellidos
                                if 'departamento' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=6, value=datos_actualizacion['departamento'])  # Columna 6: Departamento
                                if 'motivo' in datos_actualizacion:
                                    ws.cell(row=row_idx, column=7, value=datos_actualizacion['motivo'])  # Columna 7: Motivo

                                # Nota: La fecha (columna 9) NO se modifica

                                wb.save(excel_path)
                                logger.info(f"Ticket {ticket} actualizado")
                                return {"exito": True, "ticket": ticket, "mensaje": "Registro actualizado exitosamente"}

            return {"error": f"Ticket {ticket} no encontrado"}

        except Exception as e:
            logger.error(f"Error actualizando ingreso: {e}")
            return {"error": str(e)}
